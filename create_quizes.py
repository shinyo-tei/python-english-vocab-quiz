# 必要なライブラリをインポートします
from openpyxl import load_workbook # Excelファイルを操作するためのライブラリ
import numpy as np # 数値計算、特に配列のシャッフル（ランダム化）に使用

# --- 設定項目 ---
# 読み込むExcelファイルのパスを指定します
# ※このパスは実行する環境に合わせて変更する必要があります
path = './Desktop/toeic/words.xlsx'
# 読み込むExcelのシート番号を指定します
sheetnum = 2
# --- 設定項目ここまで ---

# Excelワークブック（ファイル全体）を読み込みます
wb = load_workbook(path)
# 指定したシート番号のシートを操作対象として設定します
ws = wb[f"Sheet{sheetnum}"]
# シートの最大列数と最大行数を取得します（データの範囲を把握するため）
col_count = ws.max_column
row_count = ws.max_row
# メニューとして表示する選択肢のリストを定義します
functions = ["1. 未習得単語一覧", "2. 単語テスト", "3. やめる"]

# ターミナルの文字に色をつけたり、太字にしたりするためのクラスを定義します
# ANSIエスケープシーケンスという特殊な文字列を使っています
class Color:
    BLACK          = '\033[30m'#(文字)黒
    RED            = '\033[31m'#(文字)赤
    GREEN          = '\033[32m'#(文字)緑
    YELLOW         = '\033[33m'#(文字)黄
    BLUE           = '\033[34m'#(文字)青
    MAGENTA        = '\033[35m'#(文字)マゼンタ
    CYAN           = '\033[36m'#(文字)シアン
    WHITE          = '\033[37m'#(文字)白
    COLOR_DEFAULT  = '\033[39m'#文字色をデフォルトに戻す
    BOLD           = '\033[1m'#太字
    UNDERLINE      = '\033[4m'#下線
    INVISIBLE      = '\033[08m'#不可視
    REVERCE        = '\033[07m'#文字色と背景色を反転
    BG_BLACK       = '\033[40m'#(背景)黒
    BG_RED         = '\033[41m'#(背景)赤
    BG_GREEN       = '\033[42m'#(背景)緑
    BG_YELLOW      = '\033[43m'#(背景)黄
    BG_BLUE        = '\033[44m'#(背景)青
    BG_MAGENTA     = '\033[45m'#(背景)マゼンタ
    BG_CYAN        = '\033[46m'#(背景)シアン
    BG_WHITE       = '\033[47m'#(背景)白
    BG_DEFAULT     = '\033[49m'#背景色をデフォルトに戻す
    RESET          = '\033[0m'#全てリセット

# 見出し用の装飾をする関数です（文字色と背景色を反転させます）
def heading(strings):
    return f"{Color.REVERCE}{strings}{Color.RESET}"

# 正解・不正解など、メッセージの背景に色をつけるための関数です
def render_color(color, msg):
    if color == "green":# "green"が指定されたら緑の背景色にする
        return f"{Color.BG_GREEN}{Color.WHITE}{msg}{Color.RESET}"
    elif color == "red":# "red"が指定されたら赤の背景色にする
        return f"{Color.BG_RED}{Color.WHITE}{msg}{Color.RESET}"

# 「未習得単語の一覧」を表示する関数です
def show_unfamiliar_vocab():
    print("\n" + heading("未習得単語一覧"))
    # Excelシートの1行目から最終行までループします
    for i in range(1, row_count+1):
        # 1列目の値が「0」（未習得）の単語だけを表示します
        if ws[i][0].value == 0:
            # 「行番号. 単語」の形式で表示します
            print(f"{i}. {ws[i][1].value}", end=" ")
            # 10個単語表示するごとに改行します
            if i % 10 == 0:
                print("")
    print("")

# 単語クイズを実行するメインの関数です
def do_quiz():
    # 正解した単語と間違えた単語を記録するための集合（set）を用意します
    welldone_vocab = set()
    failure_vocab = set()

    # まず、クイズ対象の単語がわかるように未習得単語一覧を表示します
    show_unfamiliar_vocab()

    # ユーザーにクイズ対象の単語番号を入力してもらいます
    word_ids = input("\n勉強する単語の番号をカンマ区切りで入力(全部を選ぶ場合はallと入力): ")
    if word_ids == "all":
        # "all"が入力されたら、習得済みかどうかにかかわらず、1から最終行までの全ての番号をリストにします
        word_ids = list(range(1, row_count+1))
    else:
        # カンマ区切りで入力された番号をリストに変換します
        word_ids = word_ids.split(",")
    # クイズの出題順をランダムにするために、単語番号のリストをシャッフルします
    np.random.shuffle(word_ids)

    # 問題文と答えの単語をペアで保存する辞書（dictionary）を作成します
    questions_dict = {}
    for row in word_ids:
        word = ws[row][1].value # 該当する行の単語を取得
        # 3列目から最後の列までループして、例文を取得します
        for col in range(2, col_count+1):
            try:
                """ 例文の中の単語を "( )" に置き換えて問題文を作成して、
                 問題文をキー、答えの単語をバリューとして辞書に保存します"""
                questions_dict[ws[row][col].value.replace(f"{word}", "( )")] = word
            except (IndexError, AttributeError):
                # 例文が空のセルなどでエラーが出た場合は、ループを抜けます
                break

    # クイズの選択肢を作成します
    options = [] # 重複しない単語のリストを作成
    for value in questions_dict.values():
        if value not in options:
            options.append(value)

    # 辞書から問題文のリストを作成し、ランダムにシャッフルします
    questions = list(questions_dict.keys())
    np.random.shuffle(questions)

    # シャッフルされた問題リストを使って、一問ずつクイズを出題します
    for question in questions:
        ans = questions_dict[question] # この問題の正解の単語
        print("\n" + heading("問題"))
        print(question)# 問題文を表示
        # 選択肢を表示します
        for idx, option in enumerate(options):
            if option == ans:
                correct_ans_idx = idx # 正解の選択肢の番号を保存
            print(f"{idx+1}. {option}", end=" ")
        print("\n")

        # ユーザーに答えの番号を入力してもらいます
        input_ans = int(input("答えの番号を入力してください: "))
        if input_ans - 1 == correct_ans_idx:
            # 入力された番号が正解と一致した場合
            print(render_color("green", "正解です！"))
            welldone_vocab.add(ans) # 正解リストに単語を追加
        else:
            # 不正解だった場合
            print(render_color("red", "不正解です。"))
            failure_vocab.add(ans)# 不正解リストに単語を追加
    
    # クイズ終了後、結果を表示します
    print("\n" + render_color("green", "あなたが正解した単語: ") + str(welldone_vocab))
    print(render_color("red", "あなたが間違えた単語: ") + str(failure_vocab))

# メニューの表示と機能の振り分けを行う関数です
def sessions():
    print("\n" + heading("メニュー"))
    for function in functions:
        print(function)
    selected_menu = int(input("メニューの番号を入力: "))

    # 入力された番号に応じて、各機能を呼び出します
    match selected_menu: #  "1. 未習得単語一覧", "2. 単語テスト", "3. やめる"
        case 1:
            show_unfamiliar_vocab()
            return 0 # プログラムを継続
        case 2:
            do_quiz()
            return 0 # プログラムを継続
        case 3:
            return 1 # プログラムを終了させるために 1 を返す

# --- ここからプログラムが開始されます ---
# 無限ループで、ユーザーが「3. やめる」を選ぶまでメニューを繰り返し表示します
while True:
    # sessions()関数の戻り値を受け取ります
    end_or_not = sessions()

    # 戻り値が 1 (終了) だったら、ループを抜けます
    if end_or_not:
        break
